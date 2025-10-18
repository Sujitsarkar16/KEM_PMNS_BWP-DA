import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import warnings
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import os

warnings.filterwarnings('ignore')

# Set style for better visualizations
plt.style.use('seaborn-v0_8')
sns.set_palette("husl")

print("="*80)
print("DATA AUDIT & SANITY CHECKS")
print("Step 1: Comprehensive Data Quality Assessment")
print("="*80)

def load_data():
    """Load the main dataset and data dictionary"""
    print("\n📊 STEP 1: LOADING DATA")
    print("-" * 50)
    
    try:
        # Load the main dataset
        df = pd.read_excel('Data/raw/IMPUTED_DATA_WITH REDUCED_columns_21_09_2025.xlsx')
        print("✅ Main dataset loaded successfully!")
        
        # Load the data dictionary
        dict_df = pd.read_excel('Data/external/PMNS F0 Mother, FAther & F1 Child Serial Final Data Dictionary (GCRF - MAS3)_20JULY2021-DEscr.xlsx')
        print("✅ Data dictionary loaded successfully!")
        
        return df, dict_df
        
    except Exception as e:
        print(f"❌ Error loading data: {e}")
        return None, None

def basic_inspection(df):
    """Perform basic data inspection"""
    print("\n📊 STEP 2: BASIC DATA INSPECTION")
    print("-" * 50)
    
    # Shape and basic info
    print(f"📏 Dataset Dimensions: {df.shape[0]:,} rows × {df.shape[1]:,} columns")
    
    # Data types
    print(f"\n📋 Data Types Summary:")
    dtype_counts = df.dtypes.value_counts()
    for dtype, count in dtype_counts.items():
        print(f"  {dtype}: {count} columns")
    
    # Memory usage
    memory_usage = df.memory_usage(deep=True).sum() / 1024**2
    print(f"\n💾 Memory Usage: {memory_usage:.2f} MB")
    
    return {
        'shape': df.shape,
        'dtypes': df.dtypes,
        'memory_usage': memory_usage
    }

def missingness_analysis(df):
    """Analyze missing values and patterns"""
    print("\n📊 STEP 3: MISSINGNESS ANALYSIS")
    print("-" * 50)
    
    # Calculate missing values
    missing_values = df.isnull().sum()
    missing_percent = (missing_values / len(df)) * 100
    
    # Create missingness summary
    missing_summary = pd.DataFrame({
        'Column': missing_values.index,
        'Missing_Count': missing_values.values,
        'Missing_Percent': missing_percent.values
    }).sort_values('Missing_Percent', ascending=False)
    
    print(f"📈 Total missing values: {missing_values.sum():,}")
    print(f"📊 Columns with missing values: {(missing_values > 0).sum()}")
    
    if missing_values.sum() > 0:
        print("\n⚠️  Columns with missing values:")
        print(missing_summary[missing_summary['Missing_Count'] > 0].head(10))
    else:
        print("\n✅ No missing values detected!")
    
    return missing_summary

def duplicate_analysis(df):
    """Check for duplicates and constant columns"""
    print("\n📊 STEP 4: DUPLICATE & CONSTANT COLUMN ANALYSIS")
    print("-" * 50)
    
    # Check for duplicate rows
    duplicate_rows = df.duplicated().sum()
    print(f"🔄 Duplicate rows: {duplicate_rows}")
    
    # Check for constant columns
    constant_cols = []
    for col in df.columns:
        if df[col].nunique() <= 1:
            constant_cols.append(col)
    
    print(f"📊 Constant columns: {len(constant_cols)}")
    if constant_cols:
        print(f"  Columns: {constant_cols}")
    
    # Check for near-constant columns (95%+ same value)
    near_constant_cols = []
    for col in df.columns:
        if df[col].dtype in ['object', 'category']:
            value_counts = df[col].value_counts()
            if len(value_counts) > 0:
                max_freq = value_counts.iloc[0] / len(df)
                if max_freq >= 0.95:
                    near_constant_cols.append((col, max_freq))
    
    print(f"📊 Near-constant columns (95%+ same value): {len(near_constant_cols)}")
    if near_constant_cols:
        for col, freq in near_constant_cols:
            print(f"  {col}: {freq:.1%} same value")
    
    return {
        'duplicate_rows': duplicate_rows,
        'constant_cols': constant_cols,
        'near_constant_cols': near_constant_cols
    }

def imputation_integrity_check(df):
    """Check for placeholder values and validate ranges"""
    print("\n📊 STEP 5: IMPUTATION INTEGRITY CHECK")
    print("-" * 50)
    
    placeholder_values = [9999, -1, -999, 'NA', 'N/A', 'NULL', 'null', 'Missing', 'missing']
    placeholder_found = {}
    
    # Check for placeholder values
    for col in df.columns:
        if df[col].dtype in ['object', 'category']:
            # Check for text placeholders
            for placeholder in ['NA', 'N/A', 'NULL', 'null', 'Missing', 'missing']:
                count = (df[col].astype(str).str.contains(placeholder, case=False, na=False)).sum()
                if count > 0:
                    placeholder_found[f"{col}_{placeholder}"] = count
        else:
            # Check for numeric placeholders
            for placeholder in [9999, -1, -999]:
                count = (df[col] == placeholder).sum()
                if count > 0:
                    placeholder_found[f"{col}_{placeholder}"] = count
    
    print(f"🔍 Placeholder values found: {len(placeholder_found)}")
    if placeholder_found:
        print("⚠️  Placeholder values detected:")
        for key, count in placeholder_found.items():
            print(f"  {key}: {count} occurrences")
    else:
        print("✅ No placeholder values detected!")
    
    # Validate numeric ranges for common health variables
    range_anomalies = {}
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    
    for col in numeric_cols:
        col_lower = col.lower()
        
        # BMI validation (if column contains 'bmi')
        if 'bmi' in col_lower:
            outliers = df[(df[col] < 10) | (df[col] > 45)][col]
            if len(outliers) > 0:
                range_anomalies[f"{col}_BMI_range"] = {
                    'min': df[col].min(),
                    'max': df[col].max(),
                    'outliers': len(outliers)
                }
        
        # Glucose validation (if column contains 'glucose' or 'gluc')
        elif 'gluc' in col_lower:
            outliers = df[df[col] > 400][col]
            if len(outliers) > 0:
                range_anomalies[f"{col}_glucose_range"] = {
                    'min': df[col].min(),
                    'max': df[col].max(),
                    'outliers': len(outliers)
                }
        
        # Age validation (if column contains 'age')
        elif 'age' in col_lower:
            outliers = df[(df[col] < 0) | (df[col] > 120)][col]
            if len(outliers) > 0:
                range_anomalies[f"{col}_age_range"] = {
                    'min': df[col].min(),
                    'max': df[col].max(),
                    'outliers': len(outliers)
                }
    
    print(f"\n📊 Range validation anomalies: {len(range_anomalies)}")
    if range_anomalies:
        print("⚠️  Range anomalies detected:")
        for key, info in range_anomalies.items():
            print(f"  {key}: range [{info['min']:.2f}, {info['max']:.2f}], {info['outliers']} outliers")
    else:
        print("✅ No range anomalies detected!")
    
    return {
        'placeholder_values': placeholder_found,
        'range_anomalies': range_anomalies
    }

def cross_reference_dictionary(df, dict_df):
    """Cross-reference variables with dictionary"""
    print("\n📊 STEP 6: DICTIONARY CROSS-REFERENCE")
    print("-" * 50)
    
    # Get column names from main dataset
    data_columns = set(df.columns)
    
    # Try to identify variable names in dictionary
    dict_columns = set()
    if not dict_df.empty:
        # Look for common column name patterns in dictionary
        for col in dict_df.columns:
            if dict_df[col].dtype == 'object':
                # Check if any values in this column match our data columns
                matching_values = set(dict_df[col].dropna().astype(str))
                dict_columns.update(matching_values.intersection(data_columns))
    
    print(f"📊 Data columns: {len(data_columns)}")
    print(f"📊 Dictionary columns: {len(dict_columns)}")
    print(f"📊 Matched columns: {len(data_columns.intersection(dict_columns))}")
    
    # Categorize variables
    numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    categorical_cols = df.select_dtypes(include=['object', 'category']).columns.tolist()
    
    print(f"\n🔢 Numerical variables: {len(numerical_cols)}")
    print(f"📝 Categorical variables: {len(categorical_cols)}")
    
    return {
        'data_columns': data_columns,
        'dict_columns': dict_columns,
        'matched_columns': data_columns.intersection(dict_columns),
        'numerical_cols': numerical_cols,
        'categorical_cols': categorical_cols
    }

def create_audit_report(df, missing_summary, duplicate_info, imputation_info, dict_info):
    """Create comprehensive audit report Excel file"""
    print("\n📊 STEP 7: CREATING AUDIT REPORT")
    print("-" * 50)
    
    # Create output directory if it doesn't exist
    output_dir = Path('Reports')
    output_dir.mkdir(exist_ok=True)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # 1. Summary sheet
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ["Data Audit Summary", ""],
        ["", ""],
        ["Dataset Shape", f"{df.shape[0]:,} rows × {df.shape[1]:,} columns"],
        ["Total Missing Values", f"{df.isnull().sum().sum():,}"],
        ["Duplicate Rows", f"{duplicate_info['duplicate_rows']:,}"],
        ["Constant Columns", f"{len(duplicate_info['constant_cols'])}"],
        ["Near-Constant Columns", f"{len(duplicate_info['near_constant_cols'])}"],
        ["Placeholder Values Found", f"{len(imputation_info['placeholder_values'])}"],
        ["Range Anomalies", f"{len(imputation_info['range_anomalies'])}"],
        ["Numerical Variables", f"{len(dict_info['numerical_cols'])}"],
        ["Categorical Variables", f"{len(dict_info['categorical_cols'])}"],
        ["", ""],
        ["Data Quality Score", f"{calculate_quality_score(df, missing_summary, duplicate_info, imputation_info):.1%}"]
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Style summary sheet
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 2. Missingness Analysis sheet
    ws_missing = wb.create_sheet("Missingness Analysis")
    missing_df = missing_summary.copy()
    missing_df = missing_df.round(2)
    
    for r in dataframe_to_rows(missing_df, index=False, header=True):
        ws_missing.append(r)
    
    # Style missingness sheet
    for cell in ws_missing[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 3. Data Types sheet
    ws_types = wb.create_sheet("Data Types")
    types_data = [
        ["Column", "Data Type", "Non-Null Count", "Null Count", "Unique Values", "Memory Usage (bytes)"]
    ]
    
    for col in df.columns:
        types_data.append([
            col,
            str(df[col].dtype),
            df[col].count(),
            df[col].isnull().sum(),
            df[col].nunique(),
            df[col].memory_usage(deep=True)
        ])
    
    for row in types_data:
        ws_types.append(row)
    
    # Style types sheet
    for cell in ws_types[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 4. Placeholder Values sheet
    if imputation_info['placeholder_values']:
        ws_placeholders = wb.create_sheet("Placeholder Values")
        placeholder_data = [["Column_Value", "Count"]]
        for key, count in imputation_info['placeholder_values'].items():
            placeholder_data.append([key, count])
        
        for row in placeholder_data:
            ws_placeholders.append(row)
        
        # Style placeholders sheet
        for cell in ws_placeholders[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # 5. Range Anomalies sheet
    if imputation_info['range_anomalies']:
        ws_ranges = wb.create_sheet("Range Anomalies")
        range_data = [["Variable", "Min Value", "Max Value", "Outlier Count"]]
        for key, info in imputation_info['range_anomalies'].items():
            range_data.append([key, info['min'], info['max'], info['outliers']])
        
        for row in range_data:
            ws_ranges.append(row)
        
        # Style ranges sheet
        for cell in ws_ranges[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
    
    # Save the workbook
    output_file = output_dir / "data_audit_report.xlsx"
    wb.save(output_file)
    print(f"✅ Audit report saved: {output_file}")
    
    return output_file

def calculate_quality_score(df, missing_summary, duplicate_info, imputation_info):
    """Calculate overall data quality score"""
    score = 100
    
    # Deduct for missing values
    missing_penalty = min(missing_summary['Missing_Percent'].max() * 2, 30)
    score -= missing_penalty
    
    # Deduct for duplicates
    duplicate_penalty = min(duplicate_info['duplicate_rows'] / len(df) * 100, 20)
    score -= duplicate_penalty
    
    # Deduct for constant columns
    constant_penalty = len(duplicate_info['constant_cols']) * 2
    score -= constant_penalty
    
    # Deduct for placeholder values
    placeholder_penalty = len(imputation_info['placeholder_values']) * 5
    score -= placeholder_penalty
    
    # Deduct for range anomalies
    range_penalty = len(imputation_info['range_anomalies']) * 3
    score -= range_penalty
    
    return max(score / 100, 0)

def create_qc_plots(df, dict_info):
    """Create quality control plots"""
    print("\n📊 STEP 8: CREATING QC PLOTS")
    print("-" * 50)
    
    # Create output directory for plots
    plots_dir = Path('PLOTS')
    plots_dir.mkdir(exist_ok=True)
    
    # 1. Missingness heatmap
    if df.isnull().sum().sum() > 0:
        plt.figure(figsize=(15, 8))
        missing_data = df.isnull()
        sns.heatmap(missing_data, cbar=True, yticklabels=False, cmap='viridis')
        plt.title('Missing Values Heatmap', fontsize=16, fontweight='bold')
        plt.xlabel('Variables', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(plots_dir / 'missingness_heatmap.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("✅ Missingness heatmap saved")
    
    # 2. Data types distribution
    plt.figure(figsize=(10, 6))
    dtype_counts = df.dtypes.value_counts()
    plt.pie(dtype_counts.values, labels=dtype_counts.index, autopct='%1.1f%%', startangle=90)
    plt.title('Data Types Distribution', fontsize=16, fontweight='bold')
    plt.axis('equal')
    plt.tight_layout()
    plt.savefig(plots_dir / 'data_types_distribution.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("✅ Data types distribution plot saved")
    
    # 3. Numerical variables histograms (limit to first 20 columns)
    numerical_cols = dict_info['numerical_cols']
    if numerical_cols:
        # Limit to first 20 columns to avoid memory issues
        cols_to_plot = numerical_cols[:20]
        n_cols = min(4, len(cols_to_plot))
        n_rows = (len(cols_to_plot) + n_cols - 1) // n_cols
        
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(20, 5*n_rows))
        if n_rows == 1:
            axes = [axes] if n_cols == 1 else axes
        else:
            axes = axes.flatten()
        
        for i, col in enumerate(cols_to_plot):
            if i < len(axes):
                df[col].hist(bins=30, ax=axes[i], alpha=0.7, edgecolor='black')
                axes[i].set_title(f'{col} Distribution', fontweight='bold')
                axes[i].set_xlabel(col)
                axes[i].set_ylabel('Frequency')
                axes[i].grid(True, alpha=0.3)
        
        # Hide empty subplots
        for i in range(len(cols_to_plot), len(axes)):
            axes[i].set_visible(False)
        
        plt.tight_layout()
        plt.savefig(plots_dir / 'numerical_histograms.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("✅ Numerical histograms saved (first 20 columns)")
    
    # 4. Correlation heatmap (for numerical variables)
    if len(numerical_cols) > 1:
        plt.figure(figsize=(15, 12))
        correlation_matrix = df[numerical_cols].corr()
        mask = np.triu(np.ones_like(correlation_matrix, dtype=bool))
        sns.heatmap(correlation_matrix, mask=mask, annot=True, cmap='coolwarm', center=0,
                   square=True, linewidths=0.5, cbar_kws={"shrink": 0.8})
        plt.title('Correlation Heatmap (Numerical Variables)', fontsize=16, fontweight='bold')
        plt.tight_layout()
        plt.savefig(plots_dir / 'correlation_heatmap.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("✅ Correlation heatmap saved")
    
    # 5. Box plots for numerical variables (limit to first 15 columns)
    if numerical_cols:
        # Limit to first 15 columns to avoid memory issues
        cols_to_plot = numerical_cols[:15]
        n_cols = min(3, len(cols_to_plot))
        n_rows = (len(cols_to_plot) + n_cols - 1) // n_cols
        
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 6*n_rows))
        if n_rows == 1:
            axes = [axes] if n_cols == 1 else axes
        else:
            axes = axes.flatten()
        
        for i, col in enumerate(cols_to_plot):
            if i < len(axes):
                df.boxplot(column=col, ax=axes[i])
                axes[i].set_title(f'{col} Box Plot', fontweight='bold')
                axes[i].grid(True, alpha=0.3)
        
        # Hide empty subplots
        for i in range(len(cols_to_plot), len(axes)):
            axes[i].set_visible(False)
        
        plt.tight_layout()
        plt.savefig(plots_dir / 'boxplots.png', dpi=300, bbox_inches='tight')
        plt.close()
        print("✅ Box plots saved (first 15 columns)")
    
    print(f"✅ All QC plots saved to {plots_dir}")

def main():
    """Main execution function"""
    # Load data
    df, dict_df = load_data()
    if df is None:
        return
    
    # Perform all analyses
    basic_info = basic_inspection(df)
    missing_summary = missingness_analysis(df)
    duplicate_info = duplicate_analysis(df)
    imputation_info = imputation_integrity_check(df)
    dict_info = cross_reference_dictionary(df, dict_df)
    
    # Create outputs
    audit_report = create_audit_report(df, missing_summary, duplicate_info, imputation_info, dict_info)
    create_qc_plots(df, dict_info)
    
    # Final summary
    print("\n" + "="*80)
    print("DATA AUDIT COMPLETE")
    print("="*80)
    print(f"✅ Audit report: {audit_report}")
    print(f"✅ QC plots: PLOTS/ directory")
    print(f"📊 Data quality score: {calculate_quality_score(df, missing_summary, duplicate_info, imputation_info):.1%}")
    print("="*80)

if __name__ == "__main__":
    main()
